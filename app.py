import os
import random
import time
from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from werkzeug.utils import secure_filename
import io

# 엑셀 처리를 위해 openpyxl 사용
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "SOME_RANDOM_SECRET_KEY"  # 세션 사용을 위한 Secret key

# 업로드 파일 저장 경로(임시). Netlify 사용 시에는 별도의 저장 방식 고려 필요.
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

################################################
# 메인 페이지
################################################
@app.route('/')
def index():
    """
    메인 페이지:
    1) 제목
    2) '명단 업로드' 버튼
    """
    return render_template('index.html')


################################################
# 파일 업로드 페이지
################################################
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """
    파일 업로드 처리:
    - GET: 업로드 폼/drag & drop 화면
    - POST: 파일 수신 후 처리
    """
    if request.method == 'POST':
        # 파일 받기
        file = request.files.get('file')
        if not file:
            flash("파일이 선택되지 않았습니다.")
            return redirect(request.url)

        filename = secure_filename(file.filename)
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(save_path)

        # 업로드 완료 -> session 에 이름 목록 저장
        names = []
        ext = os.path.splitext(filename)[1].lower()
        
        
        # # 엑셀 파일(.xlsx/.xls)이라면 openpyxl 로 열기
        # if ext in ['.xlsx', '.xls']:
        #     wb = load_workbook(save_path)
        #     ws = wb.active
        #     # 예: 첫 번째 열에 이름이 있다고 가정
        #     for row in ws.iter_rows(values_only=True):
        #         # row[0]이 실제 이름이라고 가정
        #         if row[0] is not None:
        #             names.append(str(row[0]).strip())
        
        # # 텍스트 파일(.txt)이라면 한 줄에 한 명씩
        # elif ext in ['.txt']:
        #     with open(save_path, 'r', encoding='utf-8') as f:
        #         lines = f.readlines()
        #         for line in lines:
        #             line = line.strip()
        #             if line:
        #                 names.append(line)
        
        # 엑셀 파일(.xlsx/.xls)이라면 openpyxl 로 열기
        if ext in ['.xlsx', '.xls']:
            wb = load_workbook(save_path)
            ws = wb.active
            
            # 만약 첫 줄(헤더)은 건너뛰고 싶다면 enumerate()를 이용하거나,
            # 첫 줄이 "번호", "이름"인지를 확인해서 건너뛰는 로직 추가
            is_first_row = True
            for row in ws.iter_rows(values_only=True):
                # 첫 번째 행이 '번호', '이름' 같은 헤더라면 무시
                if is_first_row:
                    is_first_row = False
                    continue
            
                # row[0] = 번호, row[1] = 이름
                if row and len(row) >= 2 and row[1] is not None:
                    names.append(str(row[1]).strip())

        elif ext in ['.txt']:
            with open(save_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
                # 첫 줄이 "번호 이름"과 같은 헤더인지, 혹은 무시해도 되는 줄인지 확인
                # 여기서는 "첫 줄은 무조건 헤더"라고 가정하고, 첫 줄을 건너뛰도록 해보겠습니다.
                is_first_line = True
            
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue  # 공백 줄이면 스킵
                
                    # 첫 줄(헤더) 건너뛰기
                    if is_first_line:
                        is_first_line = False
                        continue
                
                    # "번호 이름" 형태라면, 공백으로 split -> parts[0]=번호, parts[1]=이름
                    parts = line.split()
                
                    # 최소한 2개 이상 토큰이 있을 경우 2번째 것을 이름으로 간주
                    if len(parts) >= 2:
                        name = parts[1]
                        names.append(name)
                    else:
                        # 혹은 1개 토큰만 있을 경우 그대로 이름으로 처리
                        names.append(parts[0])

        else:
            flash("지원하지 않는 파일 형식입니다. xlsx 또는 txt 파일을 사용해주세요.")
            return redirect(request.url)

        # 세션에 저장
        session['names'] = names
        # 현재는 번호를 배정하지 않은 상태
        session['random_numbers'] = [None] * len(names)

        flash("업로드 완료!")
        return redirect(url_for('check_list'))
    else:
        return render_template('upload.html')


################################################
# 명단 확인 페이지
################################################
@app.route('/check')
def check_list():
    """
    업로드 된 명단 확인:
    - 이름 리스트
    - 번호는 아직 배정 전(blank)
    - 하단에 '돌리기' 버튼
    """
    if 'names' not in session:
        flash("명단이 없습니다. 먼저 업로드 해주세요.")
        return redirect(url_for('index'))

    names = session['names']
    random_numbers = session.get('random_numbers', [None]*len(names))
    return render_template('check.html', names=names, random_numbers=random_numbers)


################################################
# 돌리기(랜덤 순서 배정) - 로딩 애니메이션 페이지
################################################
@app.route('/shuffle')
def shuffle_numbers():
    """
    돌리기 버튼을 누르면:
    - 3,2,1...START 애니메이션(클라이언트에서 5초 지연)
    - 이후 /result 로 리다이렉트
    """
    if 'names' not in session:
        flash("명단이 없습니다. 먼저 업로드 해주세요.")
        return redirect(url_for('index'))

    # 여기서 실제 랜덤 처리
    names = session['names']
    n = len(names)
    # 1부터 n까지 랜덤 순열 생성
    random_list = list(range(1, n+1))
    random.shuffle(random_list)
    session['random_numbers'] = random_list

    # 3,2,1...START 페이지 보여주고 5초 후 /result 로 이동
    return render_template('shuffle.html')


################################################
# 결과 페이지 (랜덤 번호 매핑된 테이블 표시)
################################################
# @app.route('/result')
# def result():
#     """
#     최종 랜덤 배정된 결과를 테이블로 표시
#     - 하단에 '돌리기', '결과 저장하기' 버튼
#     """
#     if 'names' not in session or 'random_numbers' not in session:
#         flash("명단이 없습니다. 먼저 업로드 해주세요.")
#         return redirect(url_for('index'))

#     names = session['names']
#     random_numbers = session['random_numbers']

#     # 이름과 번호를 묶은 튜플 리스트
#     paired = list(zip(names, random_numbers))

#     return render_template('result.html', paired=paired)
@app.route('/result')
def result():
    if 'names' not in session or 'random_numbers' not in session:
        flash("명단이 없습니다. 먼저 업로드 해주세요.")
        return redirect(url_for('index'))

    names = session['names']
    random_numbers = session['random_numbers']

    # [{'name': n, 'number': num}, ...] 형태의 리스트
    my_list = []
    for n, num in zip(names, random_numbers):
        my_list.append({'name': n, 'number': num})

    # 이제 템플릿에 my_list를 넘겨준다.
    return render_template('result.html', my_list=my_list)

################################################
# 결과 저장하기 (Excel 다운로드)
################################################
@app.route('/download')
def download():
    """
    이름과 번호를 Excel 파일로 다운로드
    """
    if 'names' not in session or 'random_numbers' not in session:
        flash("명단이 없습니다. 먼저 업로드 해주세요.")
        return redirect(url_for('index'))

    names = session['names']
    random_numbers = session['random_numbers']

    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "결과"

    # 헤더
    ws.cell(row=1, column=1, value="No")
    ws.cell(row=1, column=2, value="이름")
    ws.cell(row=1, column=3, value="번호")

    # 데이터 작성
    for i, (name, num) in enumerate(zip(names, random_numbers), start=1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=name)
        ws.cell(row=i+1, column=3, value=num)

    # 메모리 버퍼에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 파일 다운로드 응답
    return send_file(
        output,
        as_attachment=True,
        download_name='투표결과.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)